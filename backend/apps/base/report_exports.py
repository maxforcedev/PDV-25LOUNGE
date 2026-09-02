import csv
import html
import io
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import Enum
from uuid import UUID

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

from .export_labels import export_label, export_value

FORMULA_PREFIXES = ('=', '+', '-', '@')
CURRENCY_FIELDS = ('amount', 'price', 'cost', 'value', 'revenue', 'received', 'payment', 'commission', 'discount', 'fee', 'opening', 'expected', 'informed', 'difference', 'result')


def _value(value):
    value = export_value(value)
    return json.dumps(
        _json_value(value), ensure_ascii=False, separators=(',', ':'),
    ) if isinstance(value, (dict, list, tuple)) else value


def _json_value(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return format(value, 'f')
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, Enum):
        return _json_value(value.value)
    if isinstance(value, dict):
        return {str(key): _json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_value(item) for item in value]
    return str(value)


def _csv_value(value):
    value = _value(value)
    if value is None:
        return ''
    value = str(value)
    if value.startswith(FORMULA_PREFIXES):
        try:
            Decimal(value)
        except InvalidOperation:
            value = "'" + value
    return value


def _text(value):
    """Safe text for spreadsheet metadata and ReportLab Paragraph markup."""
    return _csv_value(value)


def _paragraph(value, style):
    return Paragraph(html.escape(str(_text(value))), style)


def _company_name(branch):
    company = branch.company
    return company.trade_name or company.legal_name


def _typed_value(key, value):
    value = _value(value)
    if isinstance(value, datetime):
        return timezone.localtime(value).replace(tzinfo=None) if timezone.is_aware(value) else value
    if value is None or not isinstance(value, str):
        return value
    if key.endswith(('_at', '_date')):
        try:
            parsed = datetime.fromisoformat(value.replace('Z', '+00:00'))
            return timezone.localtime(parsed).replace(tzinfo=None) if timezone.is_aware(parsed) else parsed
        except ValueError:
            return value
    if any(part in key for part in CURRENCY_FIELDS):
        try:
            return Decimal(value)
        except InvalidOperation:
            pass
    if 'percent' in key or 'percentage' in key:
        try:
            return Decimal(value) / Decimal('100')
        except InvalidOperation:
            pass
    # OpenPyXL interprets leading '=' strings as formulas too.
    return _csv_value(value)


def _filters(request):
    return ', '.join(
        f'{export_label(key)}: {_text(value)}'
        for key, value in request.query_params.items() if key != 'export'
    ) or 'Sem filtros adicionais'


def _summary_rows(summary):
    return [
        (key, export_label(key), value)
        for key, value in (summary or {}).items()
        if not isinstance(value, (dict, list))
    ]


def _structured_sections(summary):
    sections = []
    for key, value in (summary or {}).items():
        if isinstance(value, list) and value and all(isinstance(item, dict) for item in value):
            headers = []
            for item in value:
                for item_key in item:
                    if item_key not in headers:
                        headers.append(item_key)
            if headers:
                sections.append((key, tuple(headers), value))
        elif isinstance(value, list) and value:
            sections.append((key, ('value',), [{'value': item} for item in value]))
        elif isinstance(value, dict):
            section_rows = report_key_value_rows(value, export_label(key))
            if section_rows:
                sections.append((key, ('indicator', 'value'), section_rows))
    return sections


def _sheet_title(workbook, value):
    base = ''.join(character for character in export_label(value) if character not in r'[]:*?/\\')[:31] or 'Seção'
    title = base
    suffix = 2
    while title in workbook.sheetnames:
        ending = f' {suffix}'
        title = f'{base[:31 - len(ending)]}{ending}'
        suffix += 1
    return title


def _write_sheet_table(sheet, headers, rows):
    labels = [export_label(key) for key in headers]
    sheet.append(labels)
    for row in rows:
        sheet.append([_typed_value(key, row.get(key)) for key in headers])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
    if headers:
        sheet.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{sheet.max_row}'
        sheet.freeze_panes = 'A2'
    for column, key in enumerate(headers, 1):
        values = [len(labels[column - 1]), 10]
        values.extend(len(_csv_value(row.get(key))) for row in rows)
        sheet.column_dimensions[get_column_letter(column)].width = min(max(values) + 2, 45)
        for cells in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for cell in cells:
                if isinstance(cell.value, Decimal):
                    cell.number_format = (
                        '0.00%'
                        if 'percent' in key or 'percentage' in key
                        else 'R$ #,##0.00'
                        if any(part in key for part in CURRENCY_FIELDS)
                        else '#,##0.000'
                    )
                elif isinstance(cell.value, datetime):
                    cell.number_format = 'dd/mm/yyyy hh:mm'
                elif isinstance(cell.value, date):
                    cell.number_format = 'dd/mm/yyyy'
                cell.alignment = Alignment(vertical='top', wrap_text=True)


def report_key_value_rows(data, prefix=''):
    """Flatten an already prepared response without interpreting its values."""
    rows = []
    for key, value in data.items():
        label = f'{prefix} {export_label(key)}'.strip()
        if isinstance(value, dict):
            rows.extend(report_key_value_rows(value, label))
        else:
            rows.append({'indicator': label, 'value': value})
    return rows


def render_report_export(request, *, filename, title, headers, rows, period=None, summary=None):
    """Render an already computed, permission-redacted report dataset."""
    export_format = request.query_params['export']
    headers, rows = tuple(headers), list(rows)
    filename = f'{filename.rsplit(".", 1)[0]}.{export_format}'
    labels = [export_label(key) for key in headers]
    company_name = _company_name(request.branch_context)
    metadata = (
        ('Empresa', company_name),
        ('Filial', request.branch_context.name),
        ('Período', period or ''),
        ('Filtros', _filters(request)),
    )
    if export_format == 'csv':
        output = io.StringIO(newline='')
        writer = csv.writer(output)
        writer.writerow(labels)
        writer.writerows([_csv_value(row.get(key)) for key in headers] for row in rows)
        response = HttpResponse('\ufeff' + output.getvalue(), content_type='text/csv; charset=utf-8')
    elif export_format == 'xlsx':
        output, workbook = io.BytesIO(), Workbook()
        sheet = workbook.active
        sheet.title = 'Resumo'
        for row in (['CORE PDV', _text(title)], *metadata, []):
            sheet.append([_typed_value('', value) for value in row])
        if summary:
            sheet.append(['KPIs / Resumo'])
            for key, label, value in _summary_rows(summary):
                sheet.append([_typed_value('', label), _typed_value(key, value)])
        details = workbook.create_sheet('Dados')
        _write_sheet_table(details, headers, rows)
        for key, section_headers, section_rows in _structured_sections(summary):
            section_sheet = workbook.create_sheet(_sheet_title(workbook, key))
            _write_sheet_table(section_sheet, section_headers, section_rows)
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = _csv_value(cell.value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        sheet.column_dimensions['A'].width = 34
        sheet.column_dimensions['B'].width = 55
        sheet.freeze_panes = 'A2'
        workbook.save(output)
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        output = io.BytesIO()
        page_size = portrait(A4) if len(headers) <= 5 else landscape(A4)
        document = SimpleDocTemplate(
            output,
            pagesize=page_size,
            leftMargin=.6 * cm,
            rightMargin=.6 * cm,
            topMargin=1.1 * cm,
            bottomMargin=1.1 * cm,
            allowSplitting=1,
        )
        styles = getSampleStyleSheet()
        safe_title = ParagraphStyle(
            'ReportSafeTitle',
            parent=styles['Title'],
            wordWrap='CJK',
            splitLongWords=1,
        )
        safe_normal = ParagraphStyle(
            'ReportSafeNormal',
            parent=styles['Normal'],
            wordWrap='CJK',
            splitLongWords=1,
        )
        safe_heading = ParagraphStyle(
            'ReportSafeHeading',
            parent=styles['Heading2'],
            wordWrap='CJK',
            splitLongWords=1,
        )
        safe_text = ParagraphStyle(
            'ReportSafeText',
            parent=styles['BodyText'],
            wordWrap='CJK',
            splitLongWords=1,
        )
        story = [
            _paragraph(f'CORE PDV - {title}', safe_title), Spacer(1, 4),
            _paragraph(f'Empresa: {company_name} | Filial: {request.branch_context.name}', safe_normal),
            _paragraph(f'Período: {period or ""} | Filtros: {_filters(request)} | Gerado em: {timezone.localtime():%d/%m/%Y %H:%M}', safe_normal),
        ]
        if summary:
            summary_data = [[_paragraph('Indicador', safe_text), _paragraph('Valor', safe_text)]]
            summary_data.extend(
                [_paragraph(label, safe_text), _paragraph(value, safe_text)]
                for _, label, value in _summary_rows(summary)
            )
            if len(summary_data) > 1:
                summary_table = LongTable(
                    summary_data,
                    repeatRows=1,
                    colWidths=[document.width * .55, document.width * .45],
                    splitByRow=1,
                    splitInRow=1,
                )
                summary_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('GRID', (0, 0), (-1, -1), .25, colors.grey), ('VALIGN', (0, 0), (-1, -1), 'TOP')]))
                story.extend([Spacer(1, 8), _paragraph('KPIs e resumo', safe_heading), summary_table])
        for key, section_headers, section_rows in _structured_sections(summary):
            section_labels = [export_label(item) for item in section_headers]
            section_data = [[_paragraph(label, safe_text) for label in section_labels]] + [
                [_paragraph(row.get(item), safe_text) for item in section_headers]
                for row in section_rows
            ]
            section_table = LongTable(
                section_data,
                repeatRows=1,
                colWidths=[document.width / len(section_headers)] * len(section_headers),
                splitByRow=1,
                splitInRow=1,
            )
            section_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#476A85')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('GRID', (0, 0), (-1, -1), .2, colors.grey), ('FONTSIZE', (0, 0), (-1, -1), 5), ('VALIGN', (0, 0), (-1, -1), 'TOP')]))
            story.extend([Spacer(1, 8), _paragraph(export_label(key), safe_heading), section_table])
        data = [[_paragraph(label, safe_text) for label in labels]] + [
            [_paragraph(row.get(key), safe_text) for key in headers]
            for row in rows
        ]
        if headers:
            table = LongTable(
                data,
                repeatRows=1,
                colWidths=[document.width / len(headers)] * len(headers),
                splitByRow=1,
                splitInRow=1,
            )
            table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('GRID', (0, 0), (-1, -1), .25, colors.grey), ('FONTSIZE', (0, 0), (-1, -1), 6), ('VALIGN', (0, 0), (-1, -1), 'TOP')]))
            story.extend([Spacer(1, 8), _paragraph('Dados detalhados', safe_heading), table])

        def page_number(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 7)
            canvas.drawString(.6 * cm, .55 * cm, 'CORE PDV')
            canvas.drawRightString(doc.pagesize[0] - .6 * cm, .55 * cm, f'Página {doc.page}')
            canvas.restoreState()

        document.build(story, onFirstPage=page_number, onLaterPages=page_number)
        response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
